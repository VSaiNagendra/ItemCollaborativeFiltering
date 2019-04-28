from openpyxl import load_workbook
book = load_workbook('ratingitemp.xlsx')
sheetpre = book['premiss']
sheetori = book['original']
row_count = sheetpre.max_row
column_count = sheetpre.max_column
print(row_count, column_count)
n1 = 0
n0 = 0
n2 = 0
den = 0
for i in range(0, 102):
    for j in range(0, 501):
        if sheetori.cell(row=i + 1, column=j + 1).value is not None:
            if sheetpre.cell(row=i + 1, column=j + 1).value > 0:
                den = den + 1
                if abs(sheetpre.cell(row=i + 1, column=j + 1).value - sheetori.cell(row=i + 1, column=j + 1).value) <= 1:
                    n1 += 1
                if abs(sheetpre.cell(row=i + 1, column=j + 1).value - sheetori.cell(row=i + 1, column=j + 1).value) <= 0.5:
                    n0 += 1
                if abs(sheetpre.cell(row=i + 1, column=j + 1).value - sheetori.cell(row=i + 1, column=j + 1).value) <= 1.5:
                    n2 += 1
                print(sheetpre.cell(row=i + 1, column=j + 1).value, sheetori.cell(row=i + 1, column=j + 1).value)
            break
print('Count of true positives ')
print(n2, n1, n0)
print('Total no of tested values :'+str(den))
print('Effieciency in terms of percentage :')
print(n2 / den, n1 / den, n0 / den)
